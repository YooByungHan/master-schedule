"""
Master Schedule → Excel Gantt Chart Generator  v2.0
────────────────────────────────────────────────────
사용법:
  python gantt_generator_v2.py                          # 데모 데이터로 실행
  python gantt_generator_v2.py backup.json              # JSON 백업 파일 사용
  python gantt_generator_v2.py backup.json output.xlsx  # 출력 경로 지정
  python gantt_generator_v2.py backup.json - day        # 모드 강제: day|week|month

지원하는 JSON 형식:
  · 신형: { companies:[...], projects:{...} }
  · 구형: { sections:[...], projectName:..., ... }
"""
import json, sys, os
from datetime import datetime, date, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule, FormulaRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

# ════════════════════════════════════════════════════════════
#  COLOR PALETTE
# ════════════════════════════════════════════════════════════
C = {
    # Brand
    'navy'       : '1C3A5E',
    'navy2'      : '2A5080',
    'navy3'      : '3D6FA3',
    'navy_lite'  : 'EEF6FF',
    'white'      : 'FFFFFF',
    # Text
    'text'       : '1A1816',
    'text2'      : '6B6760',
    'text3'      : 'A09B95',
    # Backgrounds
    'bg'         : 'F0EDE8',
    'bg2'        : 'F7F5F2',
    'row_odd'    : 'FFFFFF',
    'row_even'   : 'F5F3EF',
    'par_bg'     : 'DBE9FB',
    'par_bg2'    : 'C9DBEF',
    # Gantt bars
    'plan'       : '2563EB',   # blue
    'plan_lite'  : 'BFDBFE',
    'plan_dark'  : '1D4ED8',
    'act'        : '16A34A',   # green
    'act_lite'   : 'BBF7D0',
    'act_dark'   : '15803D',
    'late'       : 'DC2626',   # red (actual > plan)
    'late_lite'  : 'FECACA',
    # Section colors
    'sec1'       : '1C3A5E',
    'sec2'       : '2D6A4F',
    'sec3'       : '7B3F00',
    'sec4'       : '5B2D8E',
    # Calendar
    'wkd_sat'    : 'DBEAFE',
    'wkd_sun'    : 'FCE7F3',
    'holiday'    : 'FEF3C7',
    'today_col'  : 'FED7AA',
    'today_line' : 'EA580C',
    # Misc
    'amt_bg'     : 'FAF5FF',
    'rate_bg'    : 'FFF7ED',
    'border'     : 'D4CFC9',
    'border2'    : 'A09890',
    'green_bg'   : 'D1FAE5',
    'green_text' : '065F46',
    'yellow_bg'  : 'FEF9C3',
    'yellow_text': '92400E',
    'red_bg'     : 'FEE2E2',
    'red_text'   : '991B1B',
    'hdr_sub'    : 'C8D8EA',
}

KO_HOLIDAYS = {
    date(2025,1,1), date(2025,3,1), date(2025,5,5), date(2025,5,15),
    date(2025,6,6), date(2025,8,15), date(2025,10,3), date(2025,10,6),
    date(2025,10,7), date(2025,10,8), date(2025,10,9), date(2025,12,25),
    date(2026,1,1), date(2026,2,15), date(2026,2,16), date(2026,2,17),
    date(2026,2,18), date(2026,3,1), date(2026,5,5), date(2026,5,24),
    date(2026,6,6), date(2026,8,15), date(2026,10,3), date(2026,10,5),
    date(2026,10,6), date(2026,10,7), date(2026,10,9), date(2026,12,25),
}

# ════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ════════════════════════════════════════════════════════════
def F(hex_color):
    """PatternFill solid"""
    return PatternFill('solid', fgColor=hex_color)

def ft(bold=False, size=9, color='1A1816', name='Arial', italic=False):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def al(h='center', v='center', wrap=False, shrink=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, shrink_to_fit=shrink)

def _side(style='thin', color=None):
    color = color or C['border']
    return Side(style=style, color=color)

def bd(L=True, R=True, T=True, B=True, thick_right=False):
    r = _side('medium', C['border2']) if thick_right else _side()
    return Border(
        left  = _side() if L else Side(style=None),
        right = r       if R else Side(style=None),
        top   = _side() if T else Side(style=None),
        bottom= _side() if B else Side(style=None),
    )

def bd_div():
    """Divider border: thick right only (fixed/gantt separator)"""
    return Border(
        left  = _side(),
        right = _side('medium', '5577AA'),
        top   = _side(),
        bottom= _side(),
    )

# ════════════════════════════════════════════════════════════
#  DATE UTILITIES
# ════════════════════════════════════════════════════════════
WEEKDAY_KO = ['월','화','수','목','금','토','일']

def parse_date(s):
    if not s: return None
    try:   return datetime.strptime(s.strip(), '%Y-%m-%d').date()
    except: return None

def fmt_date(d):
    return d.strftime('%Y-%m-%d') if d else ''

def fmt_amt(v):
    try:    return f"{int(v):,}"
    except: return ''

def duration_days(s, f):
    if s and f: return max(0, (f - s).days + 1)
    return None

def is_holiday(d): return d in KO_HOLIDAYS
def is_weekend(d): return d.weekday() >= 5

def month_start(d): return d.replace(day=1)
def next_month(d):
    return (d.replace(day=28) + timedelta(4)).replace(day=1)

# ════════════════════════════════════════════════════════════
#  DATA PARSING
# ════════════════════════════════════════════════════════════
def flatten_tree(sections):
    """
    Returns flat list of dicts:
      type='sec'  → {name, color}
      type='task' → {depth, sec, id, name, name2, manager,
                     plan_s, plan_f, act_s, act_f,
                     amount, act_rate, has_ch, is_leaf}
    """
    rows = []

    def add_nodes(nodes, depth, sec_name):
        for n in nodes:
            children = n.get('ch', [])
            rows.append({
                'type'    : 'task',
                'depth'   : depth,
                'sec'     : sec_name,
                'id'      : n.get('id', ''),
                'name'    : n.get('name', ''),
                'name2'   : n.get('name2', ''),
                'manager' : n.get('manager', ''),
                'plan_s'  : parse_date(n.get('plan', {}).get('s', '')),
                'plan_f'  : parse_date(n.get('plan', {}).get('f', '')),
                'act_s'   : parse_date(n.get('actual', {}).get('s', '')),
                'act_f'   : parse_date(n.get('actual', {}).get('f', '')),
                'amount'  : n.get('amount', 0) or 0,
                'act_rate': n.get('actRate', 0) or 0,
                'has_ch'  : bool(children),
                'is_leaf' : not bool(children),
            })
            if children:
                add_nodes(children, depth + 1, sec_name)

    for sec in sections:
        rows.append({
            'type' : 'sec',
            'name' : sec.get('name', ''),
            'color': sec.get('color', C['navy']).lstrip('#'),
        })
        add_nodes(sec.get('nodes', []), 0, sec.get('name', ''))

    return rows


def calc_date_range(rows):
    dates = [
        r[k] for r in rows if r['type'] == 'task'
        for k in ('plan_s', 'plan_f', 'act_s', 'act_f') if r[k]
    ]
    if not dates:
        today = date.today()
        return date(today.year, 1, 1), date(today.year, 12, 31)
    mn = min(dates) - timedelta(7)
    mx = max(dates) + timedelta(7)
    mn = mn.replace(day=1)
    # Extend to end of that month
    mx = next_month(mx) - timedelta(1)
    return mn, mx


def decide_mode(start, end, force=None):
    if force in ('day', 'week', 'month'):
        return force
    days = (end - start).days
    if days <= 90:   return 'day'
    if days <= 365:  return 'week'
    return 'month'


def build_cols(start, end, mode):
    """
    Returns list of dicts:
      day  : {date, label_top, label_bot, is_wkd, is_hol}
      week : {date (monday), label_top, label_bot}
      month: {date (1st),   label_top, label_bot, days}
    """
    cols = []
    if mode == 'day':
        d = start
        while d <= end:
            cols.append({
                'date'     : d,
                'label_top': f"{d.year}년 {d.month}월",
                'label_bot': f"{d.day}\n{WEEKDAY_KO[d.weekday()]}",
                'is_wkd'   : is_weekend(d),
                'is_hol'   : is_holiday(d),
            })
            d += timedelta(1)
    elif mode == 'week':
        # Start from Monday before or on `start`
        d = start - timedelta(start.weekday())
        while d <= end:
            cols.append({
                'date'     : d,
                'label_top': f"{d.year}년 {d.month}월",
                'label_bot': f"W{d.isocalendar()[1]:02d}\n{d.strftime('%m/%d')}",
                'is_wkd'   : False,
                'is_hol'   : False,
            })
            d += timedelta(7)
    else:  # month
        d = month_start(start)
        while d <= end:
            nm = next_month(d)
            days_in_month = (nm - d).days
            cols.append({
                'date'     : d,
                'label_top': f"{d.year}년",
                'label_bot': f"{d.month}월",
                'is_wkd'   : False,
                'is_hol'   : False,
                'days'     : days_in_month,
            })
            d = nm
    return cols


def month_group_spans(cols):
    """Group consecutive same-label_top entries → [(label, span_count)]"""
    groups = []
    cur_lbl, cur_cnt = None, 0
    for c in cols:
        lbl = c['label_top']
        if lbl != cur_lbl:
            if cur_lbl is not None:
                groups.append((cur_lbl, cur_cnt))
            cur_lbl, cur_cnt = lbl, 1
        else:
            cur_cnt += 1
    if cur_lbl:
        groups.append((cur_lbl, cur_cnt))
    return groups

# ════════════════════════════════════════════════════════════
#  BAR DETECTION
# ════════════════════════════════════════════════════════════
def col_in_range(col_date, s, f, mode):
    """True if column date falls within [s, f]."""
    if not s or not f:
        return False
    if mode == 'day':
        return s <= col_date <= f
    elif mode == 'week':
        week_end = col_date + timedelta(6)
        return s <= week_end and f >= col_date
    else:  # month
        sm = s.year * 12 + s.month
        fm = f.year * 12 + f.month
        cm = col_date.year * 12 + col_date.month
        return sm <= cm <= fm


def bar_type(row, col_date, mode, today):
    """
    Returns one of: 'plan_act', 'act_only', 'plan_only', 'today', None
    Also returns late=True if actual is behind plan.
    """
    ps, pf = row.get('plan_s'), row.get('plan_f')
    as_, af = row.get('act_s'), row.get('act_f')
    plan_on = col_in_range(col_date, ps, pf, mode)
    act_on  = col_in_range(col_date, as_, af, mode)
    late    = bool(pf and af and af > pf)

    if plan_on and act_on:  return 'both', late
    if act_on:              return 'act',  late
    if plan_on:             return 'plan', False
    return None, False

# ════════════════════════════════════════════════════════════
#  COLUMN LAYOUT
# ════════════════════════════════════════════════════════════
FIXED = [
    ('No.',       4.0),
    ('대분류',    9.5),
    ('공 종 명',  24.0),
    ('담당자',    7.5),
    ('계획\n시작', 9.5),
    ('계획\n종료', 9.5),
    ('실적\n시작', 9.5),
    ('실적\n종료', 9.5),
    ('도급금액',  14.0),
    ('실적율',     6.5),
]
N_FIXED = len(FIXED)

GANTT_COL_W = {'day': 2.6, 'week': 6.0, 'month': 9.0}

# ════════════════════════════════════════════════════════════
#  SHEET WRITER
# ════════════════════════════════════════════════════════════
class GanttSheet:
    def __init__(self, ws, proj, co_name, rows, cols, mode, today):
        self.ws   = ws
        self.proj = proj
        self.co   = co_name
        self.rows = rows
        self.cols = cols
        self.mode = mode
        self.today= today
        self.SCL  = N_FIXED + 1   # 1-indexed start column for Gantt date cols

        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

    # ── Column widths ────────────────────────────────────
    def setup_columns(self):
        ws = self.ws
        for i, (_, w) in enumerate(FIXED, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        cw = GANTT_COL_W[self.mode]
        for i in range(len(self.cols)):
            ws.column_dimensions[get_column_letter(self.SCL + i)].width = cw

    # ── Freeze panes ─────────────────────────────────────
    def freeze(self):
        self.ws.freeze_panes = self.ws.cell(5, self.SCL)

    # ── Row heights ──────────────────────────────────────
    def set_row_heights(self, data_start_row, n_data_rows):
        ws = self.ws
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 14
        ws.row_dimensions[3].height = 19
        ws.row_dimensions[4].height = 16 if self.mode == 'day' else 14
        for r in range(data_start_row, data_start_row + n_data_rows + 2):
            ws.row_dimensions[r].height = 18

    # ── Title row (row 1) ────────────────────────────────
    def write_title(self, title):
        ws = self.ws
        last_col = N_FIXED + len(self.cols)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        c = ws.cell(1, 1, title)
        c.font  = ft(bold=True, size=14, color=C['white'])
        c.fill  = F(C['navy'])
        c.alignment = al('center', 'center')

    # ── Info row (row 2) ─────────────────────────────────
    def write_info(self):
        ws = self.ws
        p  = self.proj
        last_col = N_FIXED + len(self.cols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        today_str = self.today.strftime('%Y년 %m월 %d일')
        total_amt = fmt_amt(p.get('totalBudget', 0))
        info = (
            f"현장명: {p.get('projectName','')}   "
            f"업체: {self.co}   "
            f"작성일: {p.get('writeDate','')}   "
            f"총 도급금액: {total_amt} 원   "
            f"기준일: {today_str}"
        )
        c = ws.cell(2, 1, info)
        c.font      = ft(size=8, color=C['navy'])
        c.fill      = F(C['navy_lite'])
        c.alignment = al('left', 'center')

    # ── Fixed column headers (rows 3+4) ──────────────────
    def write_fixed_headers(self):
        ws = self.ws
        for ci, (hdr, _) in enumerate(FIXED, 1):
            ws.merge_cells(start_row=3, start_column=ci, end_row=4, end_column=ci)
            c = ws.cell(3, ci, hdr)
            c.font      = ft(bold=True, size=8.5, color=C['white'])
            c.fill      = F(C['navy'])
            c.alignment = al('center', 'center', wrap=True)
            c.border    = bd()
        # Divider on last fixed col header
        last = N_FIXED
        for r in (3, 4):
            ws.cell(r, last).border = bd_div()

    # ── Gantt month / sub headers (rows 3+4) ─────────────
    def write_gantt_headers(self):
        ws = self.ws
        cols = self.cols
        SCL  = self.SCL

        # Row 3: month groups (merged)
        groups = month_group_spans(cols)
        gc = SCL
        for lbl, span in groups:
            if span > 1:
                ws.merge_cells(start_row=3, start_column=gc, end_row=3, end_column=gc + span - 1)
            c = ws.cell(3, gc, lbl)
            c.font      = ft(bold=True, size=8, color=C['white'])
            c.fill      = F(C['navy2'])
            c.alignment = al('center', 'center')
            c.border    = bd()
            gc += span

        # Row 4: individual date labels
        for i, col in enumerate(cols):
            ci = SCL + i
            c = ws.cell(4, ci, col['label_bot'])
            c.alignment = al('center', 'center', wrap=True)
            c.border    = bd()

            is_today_col = (self.mode == 'day' and col['date'] == self.today)
            if is_today_col:
                c.fill = F(C['today_col'])
                c.font = ft(bold=True, size=7, color=C['today_line'])
            elif col['is_hol']:
                c.fill = F(C['holiday'])
                c.font = ft(size=7, color='B45309', bold=True)
            elif col['is_wkd']:
                c.fill = F(C['wkd_sat'] if col['date'].weekday()==5 else C['wkd_sun'])
                c.font = ft(size=7, color='1E3A8A')
            else:
                c.fill = F(C['hdr_sub'])
                c.font = ft(size=7, color='3D4E62')

    # ── Data rows ─────────────────────────────────────────
    def write_data(self):
        ws    = self.ws
        cols  = self.cols
        SCL   = self.SCL
        mode  = self.mode
        today = self.today

        task_no   = 0
        data_row  = 5
        even      = False

        for row in self.rows:
            ws.row_dimensions[data_row].height = 20 if row['type'] == 'sec' else 15

            if row['type'] == 'sec':
                self._write_sec_row(row, data_row)
                even = False
            else:
                task_no += 1
                even = not even
                self._write_task_row(row, data_row, task_no, even)
                self._write_gantt_cells(row, data_row, task_no, even, cols, SCL, mode, today)

            data_row += 1

        return data_row  # first empty row after data

    # ── Section row ──────────────────────────────────────
    def _write_sec_row(self, row, dr):
        ws = self.ws
        last_col = N_FIXED + len(self.cols)
        sec_clr = row.get('color', C['navy']).lstrip('#')

        ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=last_col)
        c = ws.cell(dr, 1, f"  ■  {row['name']}")
        c.font      = ft(bold=True, size=10, color=C['white'])
        c.fill      = F(sec_clr)
        c.alignment = al('left', 'center')
        c.border    = bd()

    # ── Task fixed-col row ────────────────────────────────
    def _write_task_row(self, row, dr, task_no, even):
        ws     = self.ws
        depth  = row.get('depth', 0)
        is_par = row.get('has_ch', False)
        bg     = C['par_bg'] if is_par else (C['row_even'] if even else C['row_odd'])
        indent = '  ' * (depth + 1)
        pfx    = '▶ ' if is_par else '· '
        name   = indent + pfx + row['name']

        ps, pf = row.get('plan_s'), row.get('plan_f')
        as_, af= row.get('act_s'),  row.get('act_f')
        rate   = row.get('act_rate', 0) or 0
        amt    = row.get('amount', 0) or 0
        late   = bool(pf and af and af > pf)

        dur_p  = duration_days(ps, pf)
        dur_a  = duration_days(as_, af)

        vals = [
            task_no,
            row['sec'],
            name,
            row.get('manager', ''),
            fmt_date(ps),
            fmt_date(pf),
            fmt_date(as_),
            fmt_date(af),
            int(amt) if amt else '',
            f"{rate:.1f}%" if rate else '',
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(dr, ci)
            c.border    = bd()
            c.alignment = al(
                'right' if ci in (1, 9) else 'left' if ci in (2, 3, 4) else 'center',
                'center',
                wrap=(ci == 3)
            )
            c.font = ft(
                bold   = is_par,
                size   = 8 if ci == 3 else 8.5,
                color  = C['navy'] if is_par else C['text'],
                name   = 'Arial'
            )

            # Special fills & formats
            if ci == 9:
                c.value         = int(amt) if amt else ''
                c.number_format = '#,##0'
                c.fill          = F('F3EEFF' if is_par else ('FAF5FF' if not even else 'F5EFFE'))
            elif ci == 10:
                c.value = f"{rate:.1f}%" if rate else ''
                if rate >= 100:
                    c.fill = F(C['green_bg']); c.font = ft(bold=True,size=8,color=C['green_text'])
                elif rate >= 70:
                    c.fill = F(C['yellow_bg']); c.font = ft(size=8,color=C['yellow_text'])
                elif rate > 0:
                    c.fill = F(C['red_bg'] if late else 'FFF7ED')
                    c.font = ft(size=8,color=C['red_text'] if late else 'C2410C')
                else:
                    c.fill = F('FFF7ED' if not even else 'FEF3E2')
            else:
                c.fill  = F(bg)
                c.value = val if val not in ('', 0) else ''
                if ci == 1:
                    c.value = task_no
                    c.font  = ft(size=7.5, color=C['text2'])
                    c.fill  = F('F0EDE8' if not even else 'E8E5E1')

        # Divider border on last fixed col
        ws.cell(dr, N_FIXED).border = bd_div()

    # ── Gantt bar cells ───────────────────────────────────
    def _write_gantt_cells(self, row, dr, task_no, even, cols, SCL, mode, today):
        ws = self.ws
        bg = C['row_even'] if even else C['row_odd']

        for i, col in enumerate(cols):
            ci  = SCL + i
            cd  = col['date']
            c   = ws.cell(dr, ci)

            btype, late = bar_type(row, cd, mode, today)

            # Base weekend/holiday tint
            if col['is_hol']:     base = C['holiday']
            elif col['is_wkd']:   base = C['wkd_sat'] if cd.weekday()==5 else C['wkd_sun']
            else:                 base = bg

            # Bar fill
            if btype == 'both':
                bar_clr = C['late'] if late else C['act']
            elif btype == 'act':
                bar_clr = C['late_lite'] if late else C['act_lite']
            elif btype == 'plan':
                is_par = row.get('has_ch', False)
                bar_clr = C['plan_dark'] if is_par else C['plan']
            else:
                bar_clr = base

            c.fill   = F(bar_clr)
            c.border = bd()

            # Today column: orange left border
            if mode == 'day' and cd == today:
                c.border = Border(
                    left  = Side(style='medium', color=C['today_line']),
                    right = _side(),
                    top   = _side(),
                    bottom= _side(),
                )

    # ── Legend & summary at bottom ────────────────────────
    def write_legend(self, legend_row):
        ws = self.ws
        ws.row_dimensions[legend_row].height = 16
        ws.row_dimensions[legend_row+1].height = 12

        items = [
            (C['plan'],      'PLAN (계획)'),
            (C['plan_dark'], 'PLAN (상위공종)'),
            (C['act'],       'ACTUAL (실적, 정상)'),
            (C['late'],      'ACTUAL (실적, 지연)'),
            (C['act_lite'],  '실적 진행 중'),
            (C['holiday'],   '공휴일'),
            (C['wkd_sat'],   '토요일'),
            (C['today_col'], '기준일'),
        ]

        c = ws.cell(legend_row, 1, '■ 범  례')
        c.font = ft(bold=True, size=9, color=C['white'])
        c.fill = F(C['navy'])
        c.alignment = al('center','center')
        c.border = bd()

        col = 2
        for clr, lbl in items:
            sw = ws.cell(legend_row, col)
            sw.fill   = F(clr)
            sw.border = bd()
            sw.value  = ''

            tx = ws.cell(legend_row, col + 1, lbl)
            tx.font      = ft(size=8, color=C['text2'])
            tx.alignment = al('left','center')
            tx.border    = bd()
            col += 2


# ════════════════════════════════════════════════════════════
#  STATISTICS SHEET
# ════════════════════════════════════════════════════════════
def write_stats_sheet(wb, proj, rows, today):
    ws = wb.create_sheet('통계 요약')
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    # Column widths
    col_w = [14, 28, 10, 10, 11, 11, 11, 11, 10, 10, 14, 10, 12]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_w))
    c = ws.cell(1, 1, f"공정 통계 — {proj.get('projectName','')}")
    c.font = ft(bold=True,size=13,color=C['white'])
    c.fill = F(C['navy'])
    c.alignment = al('center','center')
    ws.row_dimensions[1].height = 28

    # ── Summary KPI cards (row 3) ──
    task_rows = [r for r in rows if r['type']=='task']
    leaf_rows = [r for r in task_rows if r.get('is_leaf')]

    total_tasks  = len(task_rows)
    total_leaves = len(leaf_rows)
    total_amt    = sum(r['amount'] for r in task_rows)
    act_amt      = sum(r['amount'] * r['act_rate'] / 100 for r in task_rows)
    overall_rate = (act_amt / total_amt * 100) if total_amt else 0
    late_tasks   = [r for r in task_rows if r['plan_f'] and r['act_f'] and r['act_f'] > r['plan_f']]
    completed    = [r for r in leaf_rows  if r['act_rate'] >= 100]
    in_progress  = [r for r in leaf_rows  if 0 < r['act_rate'] < 100]
    not_started  = [r for r in leaf_rows  if r['act_rate'] == 0]

    kpis = [
        ('총 공종 수',        total_tasks,          C['navy_lite'],   C['navy']),
        ('총 도급금액',        fmt_amt(total_amt)+'원', 'F3EEFF',      '7C3AED'),
        ('실적금액',          fmt_amt(int(act_amt))+'원', C['green_bg'], C['green_text']),
        ('전체 실적율',        f"{overall_rate:.1f}%",  'FFF7ED',      'EA580C'),
        ('지연 공종',         f"{len(late_tasks)}건",  C['red_bg'],    C['red_text']),
        ('완료 공종',         f"{len(completed)}건",   C['green_bg'],  C['green_text']),
        ('진행 중',           f"{len(in_progress)}건", 'DBEAFE',       '1D4ED8'),
        ('미착수',            f"{len(not_started)}건", 'F3F4F6',       '6B7280'),
    ]

    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 10

    for ki, (lbl, val, bg, fg) in enumerate(kpis):
        col = ki + 1
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col)
        lc = ws.cell(3, col, lbl)
        lc.font      = ft(size=7.5, color=C['text2'], bold=True)
        lc.fill      = F(bg)
        lc.alignment = al('center','bottom')
        lc.border    = bd()

        vc = ws.cell(4, col, val)
        vc.font      = ft(size=10, color=fg, bold=True)
        vc.fill      = F(bg)
        vc.alignment = al('center','top')
        vc.border    = bd()

    # ── Section summary table (row 6+) ──
    ws.row_dimensions[6].height = 14
    sec_hdrs = ['대분류','총 공종수','계획금액','실적금액','실적율','완료','진행중','지연','미착수']
    for ci, h in enumerate(sec_hdrs, 1):
        c = ws.cell(6, ci, h)
        c.font = ft(bold=True, size=8.5, color=C['white'])
        c.fill = F(C['navy2'])
        c.alignment = al('center','center')
        c.border = bd()

    sec_data = defaultdict(lambda: {'tasks':0,'amt':0,'act_amt':0,'done':0,'wip':0,'late':0,'none':0})
    for r in task_rows:
        sec = r['sec']
        sec_data[sec]['tasks'] += 1
        sec_data[sec]['amt']   += r['amount']
        sec_data[sec]['act_amt'] += r['amount'] * r['act_rate'] / 100
        if r['act_rate'] >= 100:             sec_data[sec]['done'] += 1
        elif r['act_rate'] > 0:              sec_data[sec]['wip']  += 1
        elif r['plan_f'] and r['act_f'] and r['act_f'] > r['plan_f']:
                                             sec_data[sec]['late'] += 1
        else:                                sec_data[sec]['none'] += 1

    sr = 7
    for sec_name, d in sec_data.items():
        ws.row_dimensions[sr].height = 15
        rate = d['act_amt'] / d['amt'] * 100 if d['amt'] else 0
        vals = [
            sec_name, d['tasks'],
            int(d['amt']),
            int(d['act_amt']),
            f"{rate:.1f}%",
            d['done'], d['wip'], d['late'], d['none'],
        ]
        even_sr = (sr % 2 == 0)
        bg = 'F5F3EF' if even_sr else 'FFFFFF'
        for ci, v in enumerate(vals, 1):
            c = ws.cell(sr, ci, v)
            c.font = ft(size=8.5, color=C['text'])
            c.alignment = al('right' if ci in (2,3,4) else 'left' if ci==1 else 'center','center')
            c.border = bd()
            c.fill   = F(bg)
            if ci == 3:
                c.number_format = '#,##0'
            elif ci == 4:
                c.number_format = '#,##0'
            elif ci == 8 and v > 0:
                c.fill = F(C['red_bg'])
                c.font = ft(size=8.5, color=C['red_text'], bold=True)
        sr += 1

    # ── Detail table (row sr+2+) ──
    ws.row_dimensions[sr+1].height = 10
    ws.row_dimensions[sr+2].height = 15
    detail_hdrs = [
        'No.','대분류','공종명','담당자',
        '계획 시작','계획 종료','실적 시작','실적 종료',
        '기간(계획)','기간(실적)','도급금액','실적율','상태'
    ]
    for ci, h in enumerate(detail_hdrs, 1):
        c = ws.cell(sr+2, ci, h)
        c.font = ft(bold=True, size=8, color=C['white'])
        c.fill = F(C['navy'])
        c.alignment = al('center','center')
        c.border = bd()
    ws.auto_filter.ref = f"A{sr+2}:{get_column_letter(len(detail_hdrs))}{sr+2}"

    dr = sr + 3
    for task_idx, r in enumerate(task_rows, 1):
        ws.row_dimensions[dr].height = 14
        ps, pf = r.get('plan_s'), r.get('plan_f')
        as_, af= r.get('act_s'),  r.get('act_f')
        rate   = r.get('act_rate', 0)
        late   = bool(pf and af and af > pf)
        ahead  = bool(pf and af and af < pf and rate >= 100)
        if rate >= 100:      status = '✅ 완료'
        elif late:           status = '⚠️ 지연'
        elif 0 < rate < 100: status = '🔄 진행중'
        else:                status = '⏸ 미착수'

        vals = [
            task_idx, r['sec'], r['name'], r.get('manager',''),
            fmt_date(ps), fmt_date(pf), fmt_date(as_), fmt_date(af),
            duration_days(ps,pf) or '', duration_days(as_,af) or '',
            int(r['amount']) if r['amount'] else '',
            f"{rate:.1f}%" if rate else '',
            status
        ]
        even_dr = (dr % 2 == 0)
        bg = 'F5F3EF' if even_dr else 'FFFFFF'
        for ci, v in enumerate(vals, 1):
            c = ws.cell(dr, ci, v)
            c.font = ft(size=8.5)
            c.alignment = al('right' if ci in (1,9,10,11) else 'left' if ci in (2,3,4) else 'center','center')
            c.border = bd()
            if ci == 11:
                c.number_format = '#,##0'
                c.fill = F('F3EEFF' if not even_dr else 'EEE8FF')
            elif ci == 13:
                if '지연' in str(v):
                    c.fill = F(C['red_bg']); c.font = ft(size=8.5, color=C['red_text'], bold=True)
                elif '완료' in str(v):
                    c.fill = F(C['green_bg']); c.font = ft(size=8.5, color=C['green_text'])
                elif '진행' in str(v):
                    c.fill = F('DBEAFE'); c.font = ft(size=8.5, color='1D4ED8')
                else:
                    c.fill = F('F3F4F6')
            else:
                c.fill = F(bg)
        dr += 1


# ════════════════════════════════════════════════════════════
#  MAIN BUILDER
# ════════════════════════════════════════════════════════════
def build_excel(proj, co_name='', force_mode=None):
    today  = date.today()
    sections = proj.get('sections', [])
    rows   = flatten_tree(sections)

    start_d, end_d = calc_date_range(rows)
    mode = decide_mode(start_d, end_d, force_mode)
    cols = build_cols(start_d, end_d, mode)

    wb = Workbook()
    ws = wb.active
    ws.title = '간트 차트'

    g = GanttSheet(ws, proj, co_name, rows, cols, mode, today)
    g.setup_columns()
    g.freeze()

    proj_name = proj.get('projectName', 'MASTER SCHEDULE')
    g.write_title(f"▣  MASTER SCHEDULE   {proj_name}")
    g.write_info()
    g.write_fixed_headers()
    g.write_gantt_headers()

    data_last_row = g.write_data()
    g.write_legend(data_last_row + 1)
    g.set_row_heights(5, data_last_row - 5)

    # Print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = 9   # A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = '1:4'

    # Stats sheet
    write_stats_sheet(wb, proj, rows, today)

    return wb, mode


# ════════════════════════════════════════════════════════════
#  DEMO DATA
# ════════════════════════════════════════════════════════════
DEMO = {
    "companies": [{"id":"d","name":"삼성메디슨(발주)","isPrimary":True}],
    "projects": {"d": {
        "projectName": "삼성메디슨 홍천공장 증축공사",
        "companyName": "삼성메디슨(발주)",
        "writeDate": date.today().strftime('%Y-%m-%d'),
        "totalBudget": 5200000000,
        "sections": [
            {"id":"s1","name":"토목공사","color":"1c3a5e","nodes":[
                {"id":"n01","name":"부지정지 및 터파기","name2":"","manager":"김철수",
                 "plan":{"s":"2026-01-05","f":"2026-02-28"},"actual":{"s":"2026-01-05","f":"2026-02-25"},
                 "amount":120000000,"actRate":100,"ch":[]},
                {"id":"n02","name":"기초공사","name2":"말뚝+기초판","manager":"이영희",
                 "plan":{"s":"2026-02-15","f":"2026-04-30"},"actual":{"s":"2026-02-20","f":"2026-05-10"},
                 "amount":280000000,"actRate":85,"ch":[
                     {"id":"n02a","name":"PHC 파일 시공","name2":"","manager":"이영희",
                      "plan":{"s":"2026-02-15","f":"2026-03-20"},"actual":{"s":"2026-02-20","f":"2026-03-25"},
                      "amount":130000000,"actRate":100,"ch":[]},
                     {"id":"n02b","name":"기초판 콘크리트","name2":"","manager":"이영희",
                      "plan":{"s":"2026-03-21","f":"2026-04-30"},"actual":{"s":"2026-03-26","f":"2026-05-10"},
                      "amount":150000000,"actRate":70,"ch":[]},
                 ]},
                {"id":"n03","name":"지하 흙막이 공사","name2":"H-pile+토류판","manager":"박민수",
                 "plan":{"s":"2026-01-10","f":"2026-03-15"},"actual":{"s":"2026-01-10","f":"2026-03-20"},
                 "amount":90000000,"actRate":100,"ch":[]},
                {"id":"n04","name":"토목 잡공사","name2":"","manager":"박민수",
                 "plan":{"s":"2026-03-20","f":"2026-04-20"},"actual":{"s":"2026-03-22","f":""},
                 "amount":40000000,"actRate":30,"ch":[]},
            ]},
            {"id":"s2","name":"골조공사","color":"2d6a4f","nodes":[
                {"id":"n05","name":"지하층 골조","name2":"B1F~B2F","manager":"최건설",
                 "plan":{"s":"2026-03-01","f":"2026-05-31"},"actual":{"s":"2026-03-05","f":""},
                 "amount":450000000,"actRate":45,"ch":[
                     {"id":"n05a","name":"B2F 골조","name2":"","manager":"최건설",
                      "plan":{"s":"2026-03-01","f":"2026-04-15"},"actual":{"s":"2026-03-05","f":"2026-04-22"},
                      "amount":200000000,"actRate":100,"ch":[]},
                     {"id":"n05b","name":"B1F 골조","name2":"","manager":"최건설",
                      "plan":{"s":"2026-04-16","f":"2026-05-31"},"actual":{"s":"2026-04-23","f":""},
                      "amount":250000000,"actRate":15,"ch":[]},
                 ]},
                {"id":"n06","name":"지상층 골조","name2":"1F~5F","manager":"최건설",
                 "plan":{"s":"2026-06-01","f":"2026-10-31"},"actual":{"s":"","f":""},
                 "amount":680000000,"actRate":0,"ch":[
                     {"id":"n06a","name":"1~2층 골조","name2":"","manager":"최건설",
                      "plan":{"s":"2026-06-01","f":"2026-07-31"},"actual":{"s":"","f":""},
                      "amount":230000000,"actRate":0,"ch":[]},
                     {"id":"n06b","name":"3~5층 골조","name2":"","manager":"최건설",
                      "plan":{"s":"2026-08-01","f":"2026-10-31"},"actual":{"s":"","f":""},
                      "amount":450000000,"actRate":0,"ch":[]},
                 ]},
            ]},
            {"id":"s3","name":"마감공사","color":"7b3f00","nodes":[
                {"id":"n07","name":"외벽 커튼월","name2":"유리+프레임","manager":"강설비",
                 "plan":{"s":"2026-09-01","f":"2026-11-30"},"actual":{"s":"","f":""},
                 "amount":350000000,"actRate":0,"ch":[]},
                {"id":"n08","name":"내부 마감","name2":"바닥·벽·천장","manager":"정인테리어",
                 "plan":{"s":"2026-10-01","f":"2026-12-15"},"actual":{"s":"","f":""},
                 "amount":280000000,"actRate":0,"ch":[]},
                {"id":"n09","name":"옥상 방수","name2":"","manager":"강설비",
                 "plan":{"s":"2026-11-01","f":"2026-12-20"},"actual":{"s":"","f":""},
                 "amount":55000000,"actRate":0,"ch":[]},
            ]},
            {"id":"s4","name":"설비/전기공사","color":"5b2d8e","nodes":[
                {"id":"n10","name":"기계설비","name2":"공조·위생·소방","manager":"한기계",
                 "plan":{"s":"2026-07-01","f":"2026-12-15"},"actual":{"s":"","f":""},
                 "amount":520000000,"actRate":0,"ch":[
                     {"id":"n10a","name":"공조 설비","name2":"","manager":"한기계",
                      "plan":{"s":"2026-07-01","f":"2026-10-31"},"actual":{"s":"","f":""},
                      "amount":280000000,"actRate":0,"ch":[]},
                     {"id":"n10b","name":"위생·소방","name2":"","manager":"한기계",
                      "plan":{"s":"2026-09-01","f":"2026-12-15"},"actual":{"s":"","f":""},
                      "amount":240000000,"actRate":0,"ch":[]},
                 ]},
                {"id":"n11","name":"전기공사","name2":"수변전·조명·통신","manager":"전전기",
                 "plan":{"s":"2026-06-15","f":"2026-12-10"},"actual":{"s":"","f":""},
                 "amount":430000000,"actRate":0,"ch":[]},
            ]},
        ]
    }}
}

# ════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════
def main():
    args = sys.argv[1:]
    json_path   = args[0] if len(args) >= 1 and args[0] != '-' else None
    out_path    = args[1] if len(args) >= 2 and args[1] != '-' else None
    force_mode  = args[2] if len(args) >= 3 else None

    if json_path:
        with open(json_path, encoding='utf-8') as f:
            data = json.load(f)
        print(f"[info] JSON 로드: {json_path}")
    else:
        print("[demo] JSON 없음 → 샘플 데모 데이터로 생성합니다")
        data = DEMO

    # Resolve project & company
    companies = data.get('companies', [])
    projects  = data.get('projects', {})

    if companies:
        primary = next((c for c in companies if c.get('isPrimary')), companies[0])
        cid     = primary['id']
        co_name = primary['name']
        proj    = projects.get(cid, {})
    elif 'sections' in data:
        # Old-format single project JSON
        proj    = data
        co_name = data.get('companyName', '')
    else:
        cid  = next(iter(projects), None)
        proj = projects.get(cid, {}) if cid else {}
        co_name = proj.get('companyName', '')

    if not proj:
        print("ERROR: 프로젝트 데이터를 찾을 수 없습니다")
        sys.exit(1)

    wb, mode = build_excel(proj, co_name, force_mode)
    print(f"[info] 간트 모드: {mode}")

    if not out_path:
        proj_name = proj.get('projectName','공정표')
        safe = proj_name.replace('/','_').replace(' ','_').replace('(','').replace(')','')
        today_str = date.today().strftime('%Y%m%d')
        out_path = f"/mnt/user-data/outputs/간트차트_{safe}_{today_str}.xlsx"

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")
    return out_path

if __name__ == '__main__':
    main()
